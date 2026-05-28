"""
@author = allenchow
@email = 32502439@stu.hzcu.edu.cn
@current_time = 14/12/2025 02:21
"""
"""
@author = allenchow
@email = 32502439@stu.hzcu.edu.cn
@current_time = 13/12/2025 14:05
"""
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog, messagebox
import time
from tqdm import tqdm

# -------------------------- 核心配置（无需修改，已适配你的文件） --------------------------
EXCEL_FILE = "Stock_table.xlsx"  # 你的库存表路径（已定位到/mnt目录）
CHECKLIST_SHEET = "19-number"  # 核对表sheet名
STOCK_SHEET = "Stock"  # 库存主表shee9950015936522
# t名
LOG_SHEET = "log"  # 日志sheet名
# 日志表头（固定）
LOG_HEADERS = ["操作时间", "条形码", "货号", "变动量", "操作类型", "操作前库存", "操作后库存", "操作状态", "备注"]
Stock_HEADERS = ["条码/19码", "货号", "商品名称", "品牌", "尺码", "库存数量", "备注"]
CHECKLIST_HEADERS = ["条码/19码", "货号", "商品名称", "品牌", "尺码", "备注"]


def init_excel():
    """初始化Excel文件（如果不存在/缺少sheet/缺少表头，自动创建）"""

    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()

        # 核对表
        checklist = wb.active
        checklist.title = CHECKLIST_SHEET
        checklist.append(CHECKLIST_HEADERS)

        # 库存表
        stock = wb.create_sheet(STOCK_SHEET)
        stock.append(Stock_HEADERS)

        # 日志表
        log = wb.create_sheet(LOG_SHEET)
        log.append(LOG_HEADERS)

        wb.save(EXCEL_FILE)
        wb.close()
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 检查核对表
    if CHECKLIST_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(CHECKLIST_SHEET)
        sheet.append(CHECKLIST_HEADERS)

    # 检查库存表
    if STOCK_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(STOCK_SHEET)
        sheet.append(Stock_HEADERS)

    # 检查日志表
    if LOG_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(LOG_SHEET)
        sheet.append(LOG_HEADERS)

    wb.save(EXCEL_FILE)
    wb.close()


def write_log(ws, barcode, sku, change, op_type, before, after, status, remark=""):
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        barcode,
        sku,
        change,
        op_type,
        before,
        after,
        status,
        remark
    ])


# def find_in_checklist(ws, barcode):
#     for row in range(3, ws.max_row + 1):
#         if str(ws.cell(row=row, column=1).value) == barcode:
#             return {
#                 "条码/19码": ws.cell(row, 1).value,
#                 "货号": ws.cell(row, 2).value,
#                 "商品名称": ws.cell(row, 3).value,
#                 "品牌": ws.cell(row, 4).value,
#                 "尺码": ws.cell(row, 5).value,
#                 "备注": ws.cell(row, 6).value
#             }
#     return None
# 扫表模式效率低，改成内存索引

def build_index(ws, start_row):
    """
    构建 条码 -> 行号 的内存索引
    """
    index = {}
    for row in range(start_row, ws.max_row + 1):
        barcode = ws.cell(row, 1).value
        if barcode:
            index[str(barcode)] = row
    return index


def find_in_checklist(ws, checklist_index, barcode):
    row = checklist_index.get(barcode)
    if not row:
        return None

    return {
        "条码/19码": ws.cell(row, 1).value,
        "货号": ws.cell(row, 2).value,
        "商品名称": ws.cell(row, 3).value,
        "品牌": ws.cell(row, 4).value,
        "尺码": ws.cell(row, 5).value,
        "备注": ws.cell(row, 6).value
    }


# tk模式
# def manual_input(barcode):
#     root = tk.Tk()
#     root.withdraw()
#     data = {
#          "条码/19码": barcode,
#          "货号": simpledialog.askstring("录入货号", "请输入货号"),
#          "商品名称": simpledialog.askstring("录入商品名称", "请输入商品名称"),
#          "品牌": simpledialog.askstring("录入品牌", "请输入品牌"),
#          "尺码": simpledialog.askstring("录入尺码", "请输入尺码"),
#          "备注": "自建库" }
#     return data
def manual_input(barcode):
    print(f"条码 {barcode} 未在核对表中，开始建档，请依次输入信息：")

    sku = input("请输入货号: ").strip()
    if not sku:
        print("货号不能为空，建档取消")
        return None

    name = input("请输入商品名称: ").strip()
    if not name:
        print("商品名称不能为空，建档取消")
        return None

    brand = input("请输入品牌: ").strip()
    if not brand:
        print("品牌不能为空，建档取消")
        return None

    size = input("请输入尺码: ").strip()
    if not size:
        print("尺码不能为空，建档取消")
        return None

    data = {
        "条码/19码": barcode,
        "货号": sku,
        "商品名称": name,
        "品牌": brand,
        "尺码": size,
        "备注": "自建库"
    }

    print("建档完成！")
    return data


# def update_stock(ws, data):
#     for row in range(2, ws.max_row + 1):
#         if str(ws.cell(row, 1).value) == str(data["条码/19码"]):
#             before = ws.cell(row, 6).value or 0
#             ws.cell(row, 6, before + 1)
#             return before, before + 1
#
#     # 不存在，新增
#     ws.append([
#         data["条码/19码"],
#         data["货号"],
#         data["商品名称"],
#         data["品牌"],
#         data["尺码"],
#         1,
#         data["备注"]
#     ])
#     return 0, 1
# 更新
def update_stock(ws, stock_index, data):
    barcode = str(data["条码/19码"])

    if barcode in stock_index:
        row = stock_index[barcode]
        before = ws.cell(row, 6).value or 0
        ws.cell(row, 6, before + 1)
        return before, before + 1

    # 新增
    ws.append([
        data["条码/19码"],
        data["货号"],
        data["商品名称"],
        data["品牌"],
        data["尺码"],
        1,
        data["备注"]
    ])
    new_row = ws.max_row
    stock_index[barcode] = new_row
    return 0, 1


def scan_and_input(barcode):
    wb = load_workbook(EXCEL_FILE)
    checklist_ws = wb[CHECKLIST_SHEET]
    stock_ws = wb[STOCK_SHEET]
    log_ws = wb[LOG_SHEET]

    data = find_in_checklist(checklist_ws, barcode)

    if data:
        before, after = update_stock(stock_ws, data)
        write_log(log_ws, barcode, data["货号"], 1, "扫码入库", before, after, "成功")
    else:
        data = manual_input(barcode)

        # 写入 19-number
        checklist_ws.append([
            data["条码/19码"],
            data["货号"],
            data["商品名称"],
            data["品牌"],
            data["尺码"],
            data["备注"]
        ])

        before, after = update_stock(stock_ws, data)
        write_log(log_ws, barcode, data["货号"], 1, "自建入库", before, after, "成功", "自建库")

    wb.save(EXCEL_FILE)
    wb.close()
    print("录入成功！")


# def continuous_scan():
#     wb = load_workbook(EXCEL_FILE)
#     checklist_ws = wb[CHECKLIST_SHEET]
#     stock_ws = wb[STOCK_SHEET]
#     log_ws = wb[LOG_SHEET]
#
#     print("=== 连续扫码入库模式 ===")
#     print("扫码直接录入，输入 exit 退出并保存\n")
#
#     count = 0
#
#     while True:
#         barcode = input("扫码：").strip()
#
#         if not barcode:
#             continue
#
#         if barcode.lower() == "exit":
#             print("收到退出指令，正在保存数据...")
#             break
#
#         data = find_in_checklist(checklist_ws, barcode)
#
#         if data:
#             before, after = update_stock(stock_ws, data)
#             write_log(
#                 log_ws, barcode, data["货号"], 1,
#                 "扫码入库", before, after, "成功"
#             )
#         else:
#             data = manual_input(barcode)
#
#             checklist_ws.append([
#                 data["条码/19码"],
#                 data["货号"],
#                 data["商品名称"],
#                 data["品牌"],
#                 data["尺码"],
#                 data["备注"]
#             ])
#
#             before, after = update_stock(stock_ws, data)
#             write_log(
#                 log_ws, barcode, data["货号"], 1,
#                 "自建入库", before, after, "成功", "自建库"
#             )
#
#         count += 1
#         print(f"✔ 已录入 {barcode}（第 {count} 条）")
#
#     wb.save(EXCEL_FILE)
#     wb.close()
#
#     print(f"已完成，共录入 {count} 条，Excel 已保存。")

def continuous_scan():
    wb = load_workbook(EXCEL_FILE)
    checklist_ws = wb[CHECKLIST_SHEET]
    stock_ws = wb[STOCK_SHEET]
    log_ws = wb[LOG_SHEET]

    checklist_index = build_index(checklist_ws, 3)
    stock_index = build_index(stock_ws, 2)

    unknown_barcode = None

    while True:
        barcode = input("扫码：").strip()

        if barcode.lower() == "exit":
            break

        data = find_in_checklist(checklist_ws, checklist_index, barcode)

        if data:
            before, after = update_stock(stock_ws, stock_index, data)
            write_log(log_ws, barcode, data["货号"], 1,
                      "扫码入库", before, after, "成功")
        else:
            unknown_barcode = barcode
            print(f"⚠ 未知条码 {barcode}，即将进入建档流程")
            break

    wb.save(EXCEL_FILE)
    wb.close()

    return unknown_barcode


if __name__ == "__main__":
    init_excel()

    unknown = continuous_scan()

    if unknown:
        print(f"开始为条码 {unknown} 建档")
        data = manual_input(unknown)

        if data:
            wb = load_workbook(EXCEL_FILE)
            checklist_ws = wb[CHECKLIST_SHEET]
            stock_ws = wb[STOCK_SHEET]
            log_ws = wb[LOG_SHEET]

            checklist_ws.append([
                data["条码/19码"],
                data["货号"],
                data["商品名称"],
                data["品牌"],
                data["尺码"],
                data["备注"]
            ])

            before, after = update_stock(
                stock_ws,
                build_index(stock_ws, 2),
                data
            )

            write_log(log_ws, unknown, data["货号"], 1,
                      "自建入库", before, after, "成功", "自建库")

            wb.save(EXCEL_FILE)
            wb.close()

