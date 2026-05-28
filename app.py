"""
库存管理系统 v3.1 - Web可视化版
纯标准库实现，零外部依赖（仅需 openpyxl 处理 Excel）
"""

import os, re, json, sqlite3, hashlib, secrets, time
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs, urlparse, unquote
from io import BytesIO

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 用户本机使用时数据库存放在项目目录，sandbox 环境使用 /tmp
_DB_CANDIDATE = os.path.join(BASE_DIR, 'stock.db')
try:
    sqlite3.connect(_DB_CANDIDATE).close()
    DB_FILE = _DB_CANDIDATE
except Exception:
    DB_FILE = '/tmp/stock.db'
HOST = '127.0.0.1'
PORT = 5000

# ============================================================
# 数据库
# ============================================================
def get_db():
    db = sqlite3.connect(DB_FILE)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db

def row_to_dict(row):
    """sqlite3.Row -> dict"""
    if row is None:
        return None
    return dict(row)

def rows_to_list(rows):
    """[sqlite3.Row] -> [dict]"""
    return [dict(r) for r in rows] if rows else []

def init_db():
    db = sqlite3.connect(DB_FILE)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barcode TEXT UNIQUE NOT NULL,
            sku TEXT NOT NULL,
            name TEXT NOT NULL,
            brand TEXT NOT NULL DEFAULT '',
            size TEXT NOT NULL DEFAULT '',
            category TEXT NOT NULL DEFAULT '未分类',
            remark TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER UNIQUE NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barcode TEXT NOT NULL,
            sku TEXT NOT NULL DEFAULT '',
            product_name TEXT NOT NULL DEFAULT '',
            quantity_change INTEGER NOT NULL DEFAULT 0,
            operation_type TEXT NOT NULL,
            quantity_before INTEGER,
            quantity_after INTEGER,
            status TEXT NOT NULL DEFAULT '成功',
            remark TEXT NOT NULL DEFAULT '',
            operator TEXT NOT NULL DEFAULT 'admin',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'admin',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode);
        CREATE INDEX IF NOT EXISTS idx_products_sku ON products(sku);
        CREATE INDEX IF NOT EXISTS idx_products_category ON products(category);
        CREATE INDEX IF NOT EXISTS idx_products_brand ON products(brand);
        CREATE INDEX IF NOT EXISTS idx_logs_barcode ON logs(barcode);
        CREATE INDEX IF NOT EXISTS idx_logs_operation ON logs(operation_type);
        CREATE INDEX IF NOT EXISTS idx_logs_created ON logs(created_at);
    """)
    cur = db.execute("SELECT id FROM users WHERE username='admin'")
    if not cur.fetchone():
        pw = hashlib.sha256('admin123'.encode()).hexdigest()
        db.execute("INSERT INTO users (username, password, role) VALUES (?,?,?)", ('admin', pw, 'admin'))
    db.commit()
    db.close()

# ============================================================
# 工具函数
# ============================================================
def check_pw(stored, provided):
    return stored == hashlib.sha256(provided.encode()).hexdigest()

def get_product_with_stock(db, barcode):
    r = db.execute("""
        SELECT p.*, COALESCE(s.quantity, 0) as quantity
        FROM products p LEFT JOIN stock s ON p.id = s.product_id WHERE p.barcode = ?
    """, (barcode,)).fetchone()
    return dict(r) if r else None

def update_stock_qty(db, pid, change):
    r = db.execute("SELECT quantity FROM stock WHERE product_id=?", (pid,)).fetchone()
    if r:
        before = r['quantity']
        after = before + change
        db.execute("UPDATE stock SET quantity=?, updated_at=CURRENT_TIMESTAMP WHERE product_id=?", (after, pid))
    else:
        before = 0
        after = change
        db.execute("INSERT INTO stock (product_id, quantity) VALUES (?,?)", (pid, after))
    return before, after

def add_log(db, barcode, sku, name, change, op, before, after, status='成功', remark='', operator='admin'):
    db.execute(
        "INSERT INTO logs (barcode,sku,product_name,quantity_change,operation_type,quantity_before,quantity_after,status,remark,operator) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (barcode, sku, name, change, op, before, after, status, remark, operator))

def h(s):
    """HTML escape"""
    if s is None: return ''
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

# ============================================================
# 模板引擎
# ============================================================
TPL = {}

def load_templates():
    d = os.path.join(BASE_DIR, 'templates')
    if not os.path.isdir(d):
        print("templates/ 不存在，请先运行: python create_templates.py")
        return
    for f in os.listdir(d):
        if f.endswith('.html'):
            with open(os.path.join(d, f), encoding='utf-8') as fh:
                TPL[f] = fh.read()
    print(f"  已加载 {len(TPL)} 个模板")

def resolve(expr, ctx):
    """解析表达式，支持 a.b.c 和 a['key'] 和字面量"""
    expr = expr.strip()
    # 字符串字面量
    if (expr.startswith("'") and expr.endswith("'")) or (expr.startswith('"') and expr.endswith('"')):
        return expr[1:-1]
    # 数字字面量
    try:
        return int(expr)
    except:
        pass
    try:
        return float(expr)
    except:
        pass
    # 变量路径解析
    parts = []
    for seg in re.split(r'\.(?![^\[]*\])', expr):
        m = re.match(r"^(\w+)\[['\"](.+?)['\"]\]$", seg)
        if m:
            parts.append(('key', m.group(1)))
            parts.append(('attr', m.group(2)))
        else:
            parts.append(('attr', seg))
    val = ctx
    for kind, name in parts:
        if val is None:
            return ''
        if isinstance(val, dict):
            val = val.get(name, '')
        elif hasattr(val, 'keys') and name in val.keys():
            try:
                val = val[name]
            except:
                val = ''
        else:
            return ''
    return val if val is not None else ''

def eval_cond(cond_str, ctx):
    """评估条件表达式"""
    s = cond_str.strip()
    # A == B
    m = re.match(r"^(.+?)\s*==\s*(.+)$", s)
    if m:
        lv = resolve(m.group(1), ctx)
        rv = resolve(m.group(2), ctx)
        return str(lv) == str(rv)
    # A != B
    m = re.match(r"^(.+?)\s*!=\s*(.+)$", s)
    if m:
        lv = resolve(m.group(1), ctx)
        rv = resolve(m.group(2), ctx)
        return str(lv) != str(rv)
    # A > B
    m = re.match(r"^(.+?)\s*>\s*(.+)$", s)
    if m:
        lv = resolve(m.group(1), ctx)
        rv = resolve(m.group(2), ctx)
        try: return int(lv) > int(rv)
        except: return False
    # A < B
    m = re.match(r"^(.+?)\s*<\s*(.+)$", s)
    if m:
        lv = resolve(m.group(1), ctx)
        rv = resolve(m.group(2), ctx)
        try: return int(lv) < int(rv)
        except: return False
    # A in B
    m = re.match(r"^(.+?)\s+in\s+(.+)$", s)
    if m:
        item = str(resolve(m.group(1), ctx))
        lst = resolve(m.group(2), ctx)
        if isinstance(lst, (list, tuple)):
            return item in [str(x) for x in lst]
        return False
    # not A
    if s.startswith('not '):
        v = resolve(s[4:], ctx)
        return not v
    # 简单变量
    v = resolve(s, ctx)
    return bool(v) and v != 0 and v != '0'

def render_tpl(name, **ctx):
    """渲染模板（支持 extends/block/for/if/elif/else/endif/变量）"""
    if name not in TPL:
        return f"<h1>Template not found: {name}</h1>"
    src = TPL[name]
    # extends
    m = re.match(r'\s*\{%\s*extends\s+"([^"]+)"\s*%\}', src)
    if m:
        parent = TPL.get(m.group(1), '')
        child_blocks = {}
        for bm in re.finditer(r'\{%\s*block\s+(\w+)\s*%\}(.*?)\{%\s*endblock\s*%\}', src, re.DOTALL):
            child_blocks[bm.group(1)] = bm.group(2)
        def fill_block(bm):
            return child_blocks.get(bm.group(1), bm.group(2))
        src = re.sub(r'\{%\s*block\s+(\w+)\s*%\}(.*?)\{%\s*endblock\s*%\}', fill_block, parent, flags=re.DOTALL)
    return _render(src, ctx)

def _render(tpl, ctx, depth=0):
    if depth > 20: return tpl
    # {% for x in list %}...{% endfor %}
    tpl = _process_for(tpl, ctx, depth)
    # {% if %}...{% elif %}...{% else %}...{% endif %}
    tpl = _process_if(tpl, ctx, depth)
    # {{ expr }} / {{ expr | filter }}
    tpl = _process_vars(tpl, ctx)
    return tpl

def _process_for(tpl, ctx, depth):
    pat = re.compile(r'\{%\s*for\s+(\w+)\s+in\s+(\w+)\s*%\}(.*?)\{%\s*endfor\s*%\}', re.DOTALL)
    while True:
        m = pat.search(tpl)
        if not m: break
        var, lst_name, body = m.group(1), m.group(2), m.group(3)
        items = ctx.get(lst_name, [])
        parts = []
        if items:
            for i, item in enumerate(items):
                c = dict(ctx)
                c[var] = item
                c['loop'] = {'index': i+1, 'first': i==0, 'last': i==len(items)-1}
                parts.append(_render(body, c, depth+1))
        tpl = tpl[:m.start()] + ''.join(parts) + tpl[m.end():]
    return tpl

def _process_if(tpl, ctx, depth):
    # 匹配 if...elif...else...endif
    pat = re.compile(
        r'\{%\s*if\s+(.+?)\s*%\}'
        r'((?:(?!\{%\s*(?:elif|else|endif)).)*)'
        r'(\{%\s*elif\s+(.+?)\s*%\}(?:(?!\{%\s*(?:elif|else|endif)).)*)*'
        r'(\{%\s*else\s*%\}(?:(?!\{%\s*endif).)*)?'
        r'\{%\s*endif\s*%\}',
        re.DOTALL
    )
    def repl(m):
        full = m.group(0)
        # 手动解析 if/elif/else/endif 块
        blocks = re.findall(
            r'\{%\s*(if|elif)\s+(.+?)\s*%\}(.*?)(?=\{%\s*(?:elif|else|endif)\s*%\}|\{%\s*endif\s*%\})',
            full, re.DOTALL
        )
        else_m = re.search(r'\{%\s*else\s*%\}(.*?)(?=\{%\s*endif\s*%\})', full, re.DOTALL)
        else_block = else_m.group(1) if else_m else ''
        for _, cond, body in blocks:
            if eval_cond(cond, ctx):
                return _render(body, ctx, depth+1)
        return _render(else_block, ctx, depth+1) if else_block else ''
    return pat.sub(repl, tpl)

def _process_vars(tpl, ctx):
    def repl(m):
        expr = m.group(1).strip()
        # 过滤器
        filt = None
        if '|' in expr:
            expr, filt = expr.split('|', 1)
            expr, filt = expr.strip(), filt.strip()
        val = resolve(expr, ctx)
        if filt:
            if filt == 'safe':
                return str(val)
            if filt.startswith('truncate'):
                n = int(re.search(r'\d+', filt).group()) if re.search(r'\d+', filt) else 50
                val = str(val)[:n]
        return h(str(val)) if val is not None else ''
    return re.sub(r'\{\{(.*?)\}\}', repl, tpl)

# ============================================================
# HTTP 处理器
# ============================================================
class App(BaseHTTPRequestHandler):
    _sessions = {}

    def log_message(self, fmt, *a): pass

    def _sid(self):
        ck = self.headers.get('Cookie','')
        m = re.search(r'sid=([^;]+)', ck)
        return m.group(1) if m else None

    def _get_sess(self):
        s = self._sid()
        return dict(self._sessions.get(s, {}))

    def _set_sess(self, data):
        s = self._sid() or secrets.token_hex(16)
        self._sessions[s] = data
        self._my_sid = s
        return s

    def _cookie_sid(self):
        return getattr(self, '_my_sid', None) or self._sid()

    def _set_cookie(self):
        s = self._cookie_sid()
        if s:
            self.send_header('Set-Cookie', f'sid={s}; Path=/; HttpOnly')

    def _html(self, body, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self._set_cookie()
        self.end_headers()
        self.wfile.write(body.encode('utf-8'))

    def _redirect(self, url):
        self.send_response(302)
        self._set_cookie()
        self.send_header('Location', url)
        self.end_headers()

    def _json(self, data):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

    def _download(self, data, name, ctype):
        self.send_response(200)
        self.send_header('Content-Type', ctype)
        self.send_header('Content-Disposition', f'attachment; filename="{name}"')
        self.end_headers()
        self.wfile.write(data)

    def _form(self):
        ln = int(self.headers.get('Content-Length', 0))
        if not ln: return {}
        body = self.rfile.read(ln).decode('utf-8')
        return {k: v[0] if len(v)==1 else v for k, v in parse_qs(body).items()}

    def _multipart(self):
        ct = self.headers.get('Content-Type','')
        if 'multipart' not in ct: return {}, None
        boundary = ct.split('boundary=')[1].strip().strip('"')
        ln = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(ln)
        form, fdata, fname = {}, None, None
        for part in body.split(('--'+boundary).encode()):
            if not part or part.strip() in (b'', b'--'): continue
            if b'\r\n\r\n' not in part: continue
            hdr, content = part.split(b'\r\n\r\n', 1)
            if content.endswith(b'\r\n'): content = content[:-2]
            hs = hdr.decode('utf-8','replace')
            nm = re.search(r'name="([^"]+)"', hs)
            fm = re.search(r'filename="([^"]*)"', hs)
            if nm:
                if fm:
                    fname, fdata = fm.group(1), content
                else:
                    form[nm.group(1)] = content.decode('utf-8','replace')
        return form, (fname, fdata) if fdata else None

    def _need_login(self):
        if 'user_id' not in self._get_sess():
            self._redirect('/login')
            return False
        return True

    def _qp(self):
        return {k: v[0] if len(v)==1 else v for k, v in parse_qs(urlparse(self.path).query).items()}

    def _ctx(self, **extra):
        """构建模板上下文"""
        sess = self._get_sess()
        c = {
            'current_user': {'username': sess.get('username', 'admin')} if 'user_id' in sess else None,
            'request': {'endpoint': ''},
            'flash': sess.pop('flash_msg', None),
        }
        self._set_sess(sess)
        c.update(extra)
        return c

    def do_GET(self):
        path = urlparse(self.path).path.rstrip('/') or '/'
        q = self._qp()
        r = {
            '/': self.pg_dashboard,
            '/login': self.pg_login,
            '/logout': self.pg_logout,
            '/products': self.pg_products,
            '/stock/in': self.pg_stock_in,
            '/stock/out': self.pg_stock_out,
            '/stock/batch': self.pg_batch_stock,
            '/stock/manage': self.pg_stock_manage,
            '/logs': self.pg_logs,
            '/import': self.pg_import,
            '/export': self.pg_export,
        }
        if path.startswith('/products/') and path.endswith('/edit'):
            return self.pg_product_edit(path)
        if path.startswith('/api/barcode/'):
            return self.api_barcode(path)
        fn = r.get(path)
        if fn: fn(q)
        else: self._html(render_tpl('404.html', current_user=None), 404)

    def do_POST(self):
        path = urlparse(self.path).path.rstrip('/')
        r = {
            '/login': self.h_login,
            '/products/add': self.h_product_add,
            '/stock/in': self.h_stock_in,
            '/stock/in/manual': self.h_stock_in_manual,
            '/stock/out': self.h_stock_out,
            '/stock/batch': self.h_batch_stock,
            '/stock/adjust': self.h_stock_adjust,
            '/import': self.h_import,
        }
        if path.startswith('/products/') and path.endswith('/delete'):
            return self.h_product_delete(path)
        if path.startswith('/products/') and path.endswith('/edit'):
            return self.h_product_edit_post(path)
        fn = r.get(path)
        if fn: fn()
        else: self._html("<h1>404</h1>", 404)

    # ==================== 页面 ====================

    def pg_login(self, q=None):
        if 'user_id' in self._get_sess():
            return self._redirect('/')
        sess = self._get_sess()
        flash = sess.pop('flash_msg', None)
        self._set_sess(sess)
        self._html(render_tpl('login.html', current_user=None, flash=flash))

    def pg_logout(self, q=None):
        self._set_sess({'flash_msg': ('info', '已安全退出')})
        self._redirect('/login')

    def pg_dashboard(self, q=None):
        if not self._need_login(): return
        db = get_db()
        stats = {
            'total_products': db.execute("SELECT COUNT(*) FROM products").fetchone()[0],
            'total_stock': db.execute("SELECT COALESCE(SUM(quantity),0) FROM stock").fetchone()[0],
            'low_stock': db.execute("SELECT COUNT(*) FROM stock WHERE quantity>0 AND quantity<5").fetchone()[0],
            'today_in': db.execute("SELECT COUNT(*) FROM logs WHERE operation_type IN ('入库','手动入库') AND DATE(created_at)=DATE('now')").fetchone()[0],
            'today_out': db.execute("SELECT COUNT(*) FROM logs WHERE operation_type='出库' AND DATE(created_at)=DATE('now')").fetchone()[0],
            'today_scans': db.execute("SELECT COUNT(*) FROM logs WHERE DATE(created_at)=DATE('now')").fetchone()[0],
            'total_logs': db.execute("SELECT COUNT(*) FROM logs").fetchone()[0],
        }
        recent = rows_to_list(db.execute("SELECT * FROM logs ORDER BY created_at DESC LIMIT 10").fetchall())
        low_items = rows_to_list(db.execute("""
            SELECT p.barcode, p.sku, p.name, p.brand, p.size, s.quantity
            FROM products p JOIN stock s ON p.id=s.product_id
            WHERE s.quantity>0 AND s.quantity<5 ORDER BY s.quantity ASC LIMIT 10
        """).fetchall())
        cats = rows_to_list(db.execute("SELECT category, COUNT(*) as count FROM products WHERE category!='' AND category!='未分类' GROUP BY category ORDER BY count DESC LIMIT 8").fetchall())
        brands = rows_to_list(db.execute("SELECT brand, COUNT(*) as count FROM products WHERE brand!='' GROUP BY brand ORDER BY count DESC LIMIT 8").fetchall())
        db.close()
        self._html(render_tpl('dashboard.html',
            **self._ctx(stats=stats, recent_logs=recent, low_stock_items=low_items,
                        categories=cats, brands=brands)))

    def pg_products(self, q=None):
        if not self._need_login(): return
        db = get_db()
        search = q.get('search','').strip() if q else ''
        cat = q.get('category','').strip() if q else ''
        brand = q.get('brand','').strip() if q else ''
        page = int(q.get('page',1)) if q else 1
        pp = 20
        wh, pr = ['1=1'], []
        if search:
            wh.append("(p.barcode LIKE ? OR p.sku LIKE ? OR p.name LIKE ?)")
            pr += [f'%{search}%']*3
        if cat:
            wh.append("p.category=?"); pr.append(cat)
        if brand:
            wh.append("p.brand=?"); pr.append(brand)
        ws = ' AND '.join(wh)
        total = db.execute(f"SELECT COUNT(*) FROM products p WHERE {ws}", pr).fetchone()[0]
        pages = max(1, (total+pp-1)//pp)
        items = rows_to_list(db.execute(
            f"SELECT p.*, COALESCE(s.quantity,0) as quantity FROM products p LEFT JOIN stock s ON p.id=s.product_id WHERE {ws} ORDER BY p.id DESC LIMIT {pp} OFFSET {(page-1)*pp}", pr
        ).fetchall())
        cats = [r['category'] for r in db.execute("SELECT DISTINCT category FROM products WHERE category!='' ORDER BY category").fetchall()]
        brs = [r['brand'] for r in db.execute("SELECT DISTINCT brand FROM products WHERE brand!='' ORDER BY brand").fetchall()]
        db.close()
        self._html(render_tpl('products.html',
            **self._ctx(items=items, search=search, category=cat, brand=brand,
                        categories=cats, brands=brs, page=page, pages=pages, total=total)))

    def pg_product_edit(self, path):
        if not self._need_login(): return
        pid = int(path.split('/')[2])
        db = get_db()
        p = row_to_dict(db.execute("SELECT p.*, COALESCE(s.quantity,0) as quantity FROM products p LEFT JOIN stock s ON p.id=s.product_id WHERE p.id=?", (pid,)).fetchone())
        db.close()
        if not p:
            sess = self._get_sess(); sess['flash_msg'] = ('danger','产品不存在'); self._set_sess(sess)
            return self._redirect('/products')
        self._html(render_tpl('product_edit.html', **self._ctx(product=p)))

    def pg_stock_in(self, q=None):
        if not self._need_login(): return
        sess = self._get_sess()
        result = sess.pop('stock_result', None)
        self._set_sess(sess)
        self._html(render_tpl('stock_in.html', **self._ctx(result=result)))

    def pg_stock_out(self, q=None):
        if not self._need_login(): return
        sess = self._get_sess()
        result = sess.pop('stock_result', None)
        self._set_sess(sess)
        self._html(render_tpl('stock_out.html', **self._ctx(result=result)))

    def pg_stock_manage(self, q=None):
        if not self._need_login(): return
        db = get_db()
        search = q.get('search','').strip() if q else ''
        page = int(q.get('page',1)) if q else 1
        pp = 20
        wh, pr = ['1=1'], []
        if search:
            wh.append("(p.barcode LIKE ? OR p.sku LIKE ? OR p.name LIKE ?)")
            pr += [f'%{search}%']*3
        ws = ' AND '.join(wh)
        total = db.execute(f"SELECT COUNT(*) FROM products p WHERE {ws}", pr).fetchone()[0]
        pages = max(1, (total+pp-1)//pp)
        items = rows_to_list(db.execute(
            f"SELECT p.*, COALESCE(s.quantity,0) as quantity FROM products p LEFT JOIN stock s ON p.id=s.product_id WHERE {ws} ORDER BY p.id DESC LIMIT {pp} OFFSET {(page-1)*pp}", pr
        ).fetchall())
        db.close()
        self._html(render_tpl('stock_manage.html', **self._ctx(items=items, search=search, page=page, pages=pages, total=total)))

    def pg_logs(self, q=None):
        if not self._need_login(): return
        db = get_db()
        search = q.get('search','').strip() if q else ''
        op_type = q.get('op_type','').strip() if q else ''
        sd = q.get('start_date','').strip() if q else ''
        ed = q.get('end_date','').strip() if q else ''
        page = int(q.get('page',1)) if q else 1
        pp = 50
        wh, pr = ['1=1'], []
        if search:
            wh.append("(barcode LIKE ? OR sku LIKE ? OR product_name LIKE ?)")
            pr += [f'%{search}%']*3
        if op_type: wh.append("operation_type=?"); pr.append(op_type)
        if sd: wh.append("DATE(created_at)>=?"); pr.append(sd)
        if ed: wh.append("DATE(created_at)<=?"); pr.append(ed)
        ws = ' AND '.join(wh)
        total = db.execute(f"SELECT COUNT(*) FROM logs WHERE {ws}", pr).fetchone()[0]
        pages = max(1, (total+pp-1)//pp)
        items = rows_to_list(db.execute(
            f"SELECT * FROM logs WHERE {ws} ORDER BY created_at DESC LIMIT {pp} OFFSET {(page-1)*pp}", pr
        ).fetchall())
        db.close()
        self._html(render_tpl('logs.html',
            **self._ctx(items=items, search=search, op_type=op_type,
                        start_date=sd, end_date=ed, page=page, pages=pages, total=total)))

    def pg_import(self, q=None):
        if not self._need_login(): return
        self._html(render_tpl('import.html', **self._ctx()))

    def pg_export(self, q=None):
        if not self._need_login(): return
        if not HAS_OPENPYXL:
            sess = self._get_sess(); sess['flash_msg'] = ('danger','请安装 openpyxl'); self._set_sess(sess)
            return self._redirect('/')
        db = get_db()
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = '产品列表'
        ws.append(['条码/19码','货号','商品名称','品牌','尺码','分类','库存数量','备注','创建时间'])
        for p in db.execute("SELECT p.*, COALESCE(s.quantity,0) as quantity FROM products p LEFT JOIN stock s ON p.id=s.product_id ORDER BY p.id").fetchall():
            ws.append([p['barcode'],p['sku'],p['name'],p['brand'],p['size'],p['category'],p['quantity'],p['remark'],str(p['created_at'])])
        ws2 = wb.create_sheet('库存明细')
        ws2.append(['条码/19码','货号','商品名称','品牌','尺码','库存数量','最后更新'])
        for r in db.execute("SELECT p.barcode,p.sku,p.name,p.brand,p.size,s.quantity,s.updated_at FROM products p JOIN stock s ON p.id=s.product_id WHERE s.quantity>0 ORDER BY s.quantity ASC").fetchall():
            ws2.append([r['barcode'],r['sku'],r['name'],r['brand'],r['size'],r['quantity'],str(r['updated_at'])])
        ws3 = wb.create_sheet('操作日志')
        ws3.append(['操作时间','条形码','货号','商品名称','变动量','操作类型','操作前库存','操作后库存','状态','备注','操作人'])
        for l in db.execute("SELECT * FROM logs ORDER BY created_at DESC").fetchall():
            ws3.append([str(l['created_at']),l['barcode'],l['sku'],l['product_name'],l['quantity_change'],l['operation_type'],l['quantity_before'],l['quantity_after'],l['status'],l['remark'],l['operator']])
        db.close()
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        self._download(buf.getvalue(), f'库存数据导出_{ts}.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def api_barcode(self, path):
        if not self._need_login(): return
        barcode = unquote(path.split('/api/barcode/')[1])
        db = get_db()
        p = get_product_with_stock(db, barcode)
        db.close()
        if p:
            self._json({'found':True, **p})
        else:
            self._json({'found':False, 'barcode':barcode})

    # ==================== POST ====================

    def h_login(self):
        f = self._form()
        u, pw = f.get('username','').strip(), f.get('password','')
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=?", (u,)).fetchone()
        db.close()
        if user and check_pw(user['password'], pw):
            self._set_sess({'user_id': user['id'], 'username': user['username']})
            self._redirect('/')
        else:
            self._set_sess({'flash_msg': ('danger','用户名或密码错误')})
            self._redirect('/login')

    def h_product_add(self):
        if not self._need_login(): return
        f = self._form()
        bc, sku, nm = f.get('barcode','').strip(), f.get('sku','').strip(), f.get('name','').strip()
        brand, size = f.get('brand','').strip(), f.get('size','').strip()
        cat, remark = f.get('category','未分类').strip() or '未分类', f.get('remark','').strip()
        sess = self._get_sess()
        if not all([bc, sku, nm]):
            sess['flash_msg'] = ('danger','条码、货号、名称为必填项')
        else:
            db = get_db()
            if db.execute("SELECT id FROM products WHERE barcode=?", (bc,)).fetchone():
                sess['flash_msg'] = ('danger', f'条码 {bc} 已存在')
            else:
                db.execute("INSERT INTO products (barcode,sku,name,brand,size,category,remark) VALUES (?,?,?,?,?,?,?)",
                           (bc, sku, nm, brand, size, cat, remark))
                db.commit()
                sess['flash_msg'] = ('success', f'产品 {nm} 添加成功')
            db.close()
        self._set_sess(sess)
        self._redirect('/products')

    def h_product_edit_post(self, path):
        if not self._need_login(): return
        pid = int(path.split('/')[2])
        f = self._form()
        db = get_db()
        db.execute("UPDATE products SET barcode=?,sku=?,name=?,brand=?,size=?,category=?,remark=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                   (f.get('barcode',''), f.get('sku',''), f.get('name',''), f.get('brand',''),
                    f.get('size',''), f.get('category','未分类'), f.get('remark',''), pid))
        db.commit(); db.close()
        sess = self._get_sess(); sess['flash_msg'] = ('success','产品已更新'); self._set_sess(sess)
        self._redirect('/products')

    def h_product_delete(self, path):
        if not self._need_login(): return
        pid = int(path.split('/')[2])
        db = get_db()
        p = db.execute("SELECT name FROM products WHERE id=?", (pid,)).fetchone()
        if p:
            db.execute("DELETE FROM products WHERE id=?", (pid,)); db.commit()
            sess = self._get_sess(); sess['flash_msg'] = ('success', f'{p["name"]} 已删除'); self._set_sess(sess)
        db.close()
        self._redirect('/products')

    def h_stock_in(self):
        if not self._need_login(): return
        f = self._form()
        bc = f.get('barcode','').strip()
        qty = int(f.get('quantity',1))
        sess = self._get_sess()
        if not bc:
            sess['flash_msg'] = ('danger','请输入条码')
        elif qty < 1:
            sess['flash_msg'] = ('danger','数量必须大于0')
        else:
            db = get_db()
            p = get_product_with_stock(db, bc)
            if p:
                before, after = update_stock_qty(db, p['id'], qty)
                add_log(db, bc, p['sku'], p['name'], qty, '入库', before, after)
                db.commit()
                sess['stock_result'] = {'success':True, 'barcode':bc, 'sku':p['sku'], 'name':p['name'], 'qty':qty, 'before':before, 'after':after}
                sess['flash_msg'] = ('success', f'{bc} 入库 {qty} 件（{before}→{after}）')
            else:
                sess['stock_result'] = {'success':False, 'barcode':bc, 'unknown':True}
                sess['flash_msg'] = ('warning', f'条码 {bc} 未知，请先建档')
            db.close()
        self._set_sess(sess)
        self._redirect('/stock/in')

    def h_stock_in_manual(self):
        if not self._need_login(): return
        f = self._form()
        bc, sku, nm = f.get('barcode','').strip(), f.get('sku','').strip(), f.get('name','').strip()
        brand, size, qty = f.get('brand','').strip(), f.get('size','').strip(), int(f.get('quantity',1))
        sess = self._get_sess()
        if not all([bc, sku, nm]):
            sess['flash_msg'] = ('danger','条码、货号、名称为必填项')
        else:
            db = get_db()
            ex = db.execute("SELECT id FROM products WHERE barcode=?", (bc,)).fetchone()
            pid = ex['id'] if ex else db.execute("INSERT INTO products (barcode,sku,name,brand,size,remark) VALUES (?,?,?,?,?,?)",
                                                  (bc,sku,nm,brand,size,'手动建档')).lastrowid
            before, after = update_stock_qty(db, pid, qty)
            add_log(db, bc, sku, nm, qty, '手动入库', before, after, '成功', '手动建档入库')
            db.commit(); db.close()
            sess['flash_msg'] = ('success', f'{bc} 建档并入库 {qty} 件')
        self._set_sess(sess)
        self._redirect('/stock/in')

    def h_stock_out(self):
        if not self._need_login(): return
        f = self._form()
        bc = f.get('barcode','').strip()
        qty = int(f.get('quantity',1))
        sess = self._get_sess()
        if not bc:
            sess['flash_msg'] = ('danger','请输入条码')
        elif qty < 1:
            sess['flash_msg'] = ('danger','数量必须大于0')
        else:
            db = get_db()
            p = get_product_with_stock(db, bc)
            if p:
                if p['quantity'] < qty:
                    sess['flash_msg'] = ('danger', f'库存不足！当前 {p["quantity"]}，无法出库 {qty}')
                else:
                    before, after = update_stock_qty(db, p['id'], -qty)
                    add_log(db, bc, p['sku'], p['name'], -qty, '出库', before, after)
                    db.commit()
                    sess['stock_result'] = {'success':True, 'barcode':bc, 'sku':p['sku'], 'name':p['name'], 'qty':qty, 'before':before, 'after':after}
                    sess['flash_msg'] = ('success', f'{bc} 出库 {qty} 件（{before}→{after}）')
            else:
                sess['flash_msg'] = ('danger', f'条码 {bc} 不存在')
            db.close()
        self._set_sess(sess)
        self._redirect('/stock/out')

    def pg_batch_stock(self, q=None):
        if not self._need_login(): return
        sess = self._get_sess()
        result = sess.pop('batch_result', None)
        self._set_sess(sess)
        self._html(render_tpl('batch_stock.html', **self._ctx(result=result)))

    def h_batch_stock(self):
        if not self._need_login(): return
        f = self._form()
        mode = f.get('mode', 'in')  # 'in' or 'out'
        barcodes_text = f.get('barcodes', '').strip()
        default_qty = int(f.get('default_qty', 1))

        sess = self._get_sess()
        if not barcodes_text:
            sess['flash_msg'] = ('danger', '请输入条码')
            self._set_sess(sess)
            return self._redirect('/stock/batch')

        # 解析条码：每行一个，支持 "条码 数量" 格式（空格分隔）
        lines = [l.strip() for l in barcodes_text.split('\n') if l.strip()]
        results = []
        success_count = 0
        fail_count = 0
        db = get_db()

        for line in lines:
            parts = line.split()
            barcode = parts[0]
            qty = int(parts[1]) if len(parts) > 1 else default_qty

            p = get_product_with_stock(db, barcode)
            if not p:
                results.append({'barcode': barcode, 'status': 'fail', 'msg': '条码不存在'})
                fail_count += 1
                continue

            if mode == 'in':
                before, after = update_stock_qty(db, p['id'], qty)
                add_log(db, barcode, p['sku'], p['name'], qty, '批量入库', before, after)
                results.append({'barcode': barcode, 'sku': p['sku'], 'name': p['name'],
                                'qty': qty, 'before': before, 'after': after, 'status': 'ok'})
                success_count += 1
            else:  # out
                if p['quantity'] < qty:
                    results.append({'barcode': barcode, 'sku': p['sku'], 'name': p['name'],
                                    'status': 'fail', 'msg': f'库存不足({p["quantity"]})'})
                    fail_count += 1
                    continue
                before, after = update_stock_qty(db, p['id'], -qty)
                add_log(db, barcode, p['sku'], p['name'], -qty, '批量出库', before, after)
                results.append({'barcode': barcode, 'sku': p['sku'], 'name': p['name'],
                                'qty': qty, 'before': before, 'after': after, 'status': 'ok'})
                success_count += 1

        db.commit()
        db.close()

        sess['batch_result'] = {
            'mode': mode,
            'results': results,
            'success': success_count,
            'fail': fail_count,
            'total': len(results)
        }
        op_name = '入库' if mode == 'in' else '出库'
        sess['flash_msg'] = ('success' if fail_count == 0 else 'warning',
                             f'批量{op_name}完成：成功 {success_count}，失败 {fail_count}')
        self._set_sess(sess)
        self._redirect('/stock/batch')

    def h_stock_adjust(self):
        if not self._need_login(): return
        f = self._form()
        pid, new_qty = int(f.get('product_id',0)), int(f.get('quantity',0))
        remark = f.get('remark','').strip()
        sess = self._get_sess()
        db = get_db()
        p = db.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
        if not p:
            sess['flash_msg'] = ('danger','产品不存在')
        elif new_qty < 0:
            sess['flash_msg'] = ('danger','库存不能为负')
        else:
            st = db.execute("SELECT quantity FROM stock WHERE product_id=?", (pid,)).fetchone()
            before = st['quantity'] if st else 0
            if st:
                db.execute("UPDATE stock SET quantity=?,updated_at=CURRENT_TIMESTAMP WHERE product_id=?", (new_qty, pid))
            else:
                db.execute("INSERT INTO stock (product_id,quantity) VALUES (?,?)", (pid, new_qty))
            add_log(db, p['barcode'], p['sku'], p['name'], new_qty-before, '库存调整', before, new_qty, '成功', remark or '手动调整')
            db.commit()
            sess['flash_msg'] = ('success', f'{p["name"]} 库存 {before}→{new_qty}')
        db.close()
        self._set_sess(sess)
        self._redirect('/stock/manage')

    def h_import(self):
        if not self._need_login(): return
        sess = self._get_sess()
        if not HAS_OPENPYXL:
            sess['flash_msg'] = ('danger','请安装 openpyxl')
            self._set_sess(sess); return self._redirect('/import')
        form, file_info = self._multipart()
        if not file_info or not file_info[1]:
            sess['flash_msg'] = ('danger','请选择文件')
            self._set_sess(sess); return self._redirect('/import')
        fname, fdata = file_info
        if not fname.endswith(('.xlsx','.xls')):
            sess['flash_msg'] = ('danger','仅支持 .xlsx/.xls')
            self._set_sess(sess); return self._redirect('/import')
        try:
            wb = openpyxl.load_workbook(BytesIO(fdata))
            db = get_db()
            added = updated = 0
            for sn in wb.sheetnames:
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: continue
                hdrs = [str(h).strip() if h else '' for h in rows[0]]
                for row in rows[1:]:
                    if not row or not any(row): continue
                    rd = dict(zip(hdrs, row)) if any(hdrs) else {}
                    bc = str(rd.get('条码/19码') or rd.get('条形码') or rd.get('ISBN码') or (row[0] if row else '')).strip()
                    if not bc: continue
                    sku = str(rd.get('货号','') or (row[1] if len(row)>1 else '')).strip()
                    nm = str(rd.get('商品名称','') or (row[2] if len(row)>2 else '')).strip()
                    brand = str(rd.get('品牌','') or (row[3] if len(row)>3 else '')).strip()
                    size = str(rd.get('尺码','') or (row[4] if len(row)>4 else '')).strip()
                    remark = str(rd.get('备注','') or (row[5] if len(row)>5 else '')).strip()
                    if not sku and not nm: continue
                    ex = db.execute("SELECT id FROM products WHERE barcode=?", (bc,)).fetchone()
                    if ex:
                        db.execute("UPDATE products SET sku=CASE WHEN ?='' THEN sku ELSE ? END, name=CASE WHEN ?='' THEN name ELSE ? END, brand=CASE WHEN ?='' THEN brand ELSE ? END, size=CASE WHEN ?='' THEN size ELSE ? END, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                                   (sku,sku,nm,nm,brand,brand,size,size,ex['id']))
                        pid = ex['id']; updated += 1
                    else:
                        cur = db.execute("INSERT INTO products (barcode,sku,name,brand,size,remark) VALUES (?,?,?,?,?,?)", (bc,sku,nm,brand,size,remark))
                        pid = cur.lastrowid; added += 1
                    qty = rd.get('库存数量', row[5] if len(row)>5 else None)
                    if qty is not None:
                        try:
                            q = int(float(str(qty)))
                            st = db.execute("SELECT quantity FROM stock WHERE product_id=?", (pid,)).fetchone()
                            if st:
                                db.execute("UPDATE stock SET quantity=?,updated_at=CURRENT_TIMESTAMP WHERE product_id=?", (q, pid))
                            else:
                                db.execute("INSERT INTO stock (product_id,quantity) VALUES (?,?)", (pid, q))
                        except: pass
            db.commit(); db.close(); wb.close()
            sess['flash_msg'] = ('success', f'导入完成！新增 {added}，更新 {updated}')
        except Exception as e:
            sess['flash_msg'] = ('danger', f'导入失败：{e}')
        self._set_sess(sess)
        self._redirect('/import')

# ============================================================
# 启动
# ============================================================
if __name__ == '__main__':
    init_db()
    load_templates()
    if not TPL:
        print("请先运行: python create_templates.py"); exit(1)
    print(f"\n{'='*50}")
    print(f"  库存管理系统 v3.1")
    print(f"  http://{HOST}:{PORT}")
    print(f"  账号: admin / admin123")
    print(f"{'='*50}\n")
    server = HTTPServer((HOST, PORT), App)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止"); server.server_close()
