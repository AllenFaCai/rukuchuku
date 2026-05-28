"""
Excel 数据导入工具
将现有 Excel 文件中的数据导入到 SQLite 数据库中
支持旧版库存表.xlsx 和 Stock_table.xlsx 两种格式
"""

import sqlite3
import os
import sys

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

_DB_DIR = os.path.dirname(os.path.abspath(__file__))
_DB_CANDIDATE = os.path.join(_DB_DIR, 'stock.db')
try:
    sqlite3.connect(_DB_CANDIDATE).close()
    DB_FILE = _DB_CANDIDATE
except Exception:
    DB_FILE = '/tmp/stock.db'


def get_db():
    db = sqlite3.connect(DB_FILE)
    db.row_factory = sqlite3.Row
    return db


def init_db():
    """初始化数据库"""
    db = sqlite3.connect(DB_FILE)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barcode TEXT UNIQUE NOT NULL,
            sku TEXT NOT NULL,
            name TEXT NOT NULL,
            brand TEXT NOT NULL DEFAULT '',
            size TEXT NOT NULL DEFAULT '',
            category TEXT NOT NULL DEFAULT '未分类',
            remark TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER UNIQUE NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barcode TEXT NOT NULL,
            sku TEXT NOT NULL DEFAULT '',
            product_name TEXT NOT NULL DEFAULT '',
            quantity_change INTEGER NOT NULL DEFAULT 0,
            operation_type TEXT NOT NULL,
            quantity_before INTEGER,
            quantity_after INTEGER,
            status TEXT NOT NULL DEFAULT '成功',
            remark TEXT NOT NULL DEFAULT '',
            operator TEXT NOT NULL DEFAULT 'system',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'admin',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode);
        CREATE INDEX IF NOT EXISTS idx_products_sku ON products(sku);
        CREATE INDEX IF NOT EXISTS idx_logs_barcode ON logs(barcode);
        CREATE INDEX IF NOT EXISTS idx_logs_created ON logs(created_at);
    """)
    db.commit()
    db.close()
    print("数据库初始化完成")


def import_stock_table(filepath):
    """
    导入 Stock_table.xlsx（新版格式）
    Sheet: 19-number, Stock, log
    """
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return

    wb = openpyxl.load_workbook(filepath)
    db = get_db()
    added = 0
    updated = 0

    # 导入 19-number (核对表) -> products
    if '19-number' in wb.sheetnames:
        ws = wb['19-number']
        print(f"\n导入核对表 (19-number)...")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if row_idx <= 2 or not row or not row[0]:  # 跳过前两行（可能有标题行）
                continue
            barcode = str(row[0]).strip() if row[0] else ''
            sku = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            brand = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            size = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            remark = str(row[5]).strip() if len(row) > 5 and row[5] else ''

            if not barcode or not sku:
                continue

            existing = db.execute("SELECT id FROM products WHERE barcode=?", (barcode,)).fetchone()
            if existing:
                db.execute(
                    "UPDATE products SET sku=?, name=?, brand=?, size=?, remark=?, "
                    "updated_at=CURRENT_TIMESTAMP WHERE barcode=?",
                    (sku, name, brand, size, remark, barcode)
                )
                updated += 1
            else:
                db.execute(
                    "INSERT INTO products (barcode, sku, name, brand, size, remark) VALUES (?,?,?,?,?,?)",
                    (barcode, sku, name, brand, size, remark)
                )
                added += 1
        print(f"  核对表: 新增 {added}, 更新 {updated}")

    # 导入 Stock (库存表) -> stock
    if 'Stock' in wb.sheetnames:
        ws = wb['Stock']
        print(f"\n导入库存表 (Stock)...")
        stock_added = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if row_idx <= 1 or not row or not row[0]:  # 跳过表头
                continue
            barcode = str(row[0]).strip() if row[0] else ''
            qty = row[5] if len(row) > 5 else 0

            if not barcode:
                continue

            try:
                qty = int(float(str(qty))) if qty else 0
            except (ValueError, TypeError):
                qty = 0

            product = db.execute("SELECT id FROM products WHERE barcode=?", (barcode,)).fetchone()
            if product:
                stock = db.execute("SELECT id FROM stock WHERE product_id=?", (product['id'],)).fetchone()
                if stock:
                    db.execute("UPDATE stock SET quantity=?, updated_at=CURRENT_TIMESTAMP WHERE product_id=?",
                               (qty, product['id']))
                else:
                    db.execute("INSERT INTO stock (product_id, quantity) VALUES (?, ?)",
                               (product['id'], qty))
                stock_added += 1
        print(f"  库存表: 导入 {stock_added} 条库存记录")

    # 导入 log (日志表) -> logs
    if 'log' in wb.sheetnames:
        ws = wb['log']
        print(f"\n导入日志表 (log)...")
        log_added = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if row_idx <= 1 or not row or not row[0]:  # 跳过表头
                continue
            try:
                db.execute(
                    "INSERT INTO logs (created_at, barcode, sku, quantity_change, operation_type, "
                    "quantity_before, quantity_after, status, remark) VALUES (?,?,?,?,?,?,?,?,?)",
                    (str(row[0]) if row[0] else '', str(row[1]) if len(row) > 1 else '',
                     str(row[2]) if len(row) > 2 else '',
                     int(row[3]) if len(row) > 3 and row[3] else 0,
                     str(row[4]) if len(row) > 4 else '',
                     int(row[5]) if len(row) > 5 and row[5] else None,
                     int(row[6]) if len(row) > 6 and row[6] else None,
                     str(row[7]) if len(row) > 7 else '成功',
                     str(row[8]) if len(row) > 8 else '')
                )
                log_added += 1
            except Exception:
                continue
        print(f"  日志表: 导入 {log_added} 条日志记录")

    db.commit()
    db.close()
    wb.close()
    print(f"\n导入完成！总计: 新增 {added} 个产品, 更新 {updated} 个产品")


def import_old_stock(filepath):
    """
    导入 库存表.xlsx（旧版格式）
    Sheet: ISBN-货号对照表, 操作日志
    """
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return

    wb = openpyxl.load_workbook(filepath)
    db = get_db()
    added = 0

    # 导入 ISBN-货号对照表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n导入 Sheet: {sheet_name}")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if row_idx <= 1 or not row or not row[0]:
                continue
            barcode = str(row[0]).strip() if row[0] else ''
            sku = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            qty = row[2] if len(row) > 2 else 0

            if not barcode or not sku:
                continue

            try:
                qty = int(float(str(qty))) if qty else 0
            except (ValueError, TypeError):
                qty = 0

            existing = db.execute("SELECT id FROM products WHERE barcode=?", (barcode,)).fetchone()
            if not existing:
                cur = db.execute(
                    "INSERT INTO products (barcode, sku, name, remark) VALUES (?,?,?,?)",
                    (barcode, sku, sku, '旧版导入')
                )
                db.execute("INSERT INTO stock (product_id, quantity) VALUES (?, ?)",
                           (cur.lastrowid, qty))
                added += 1

    db.commit()
    db.close()
    wb.close()
    print(f"\n导入完成！新增 {added} 个产品")


if __name__ == '__main__':
    print("=" * 50)
    print("  Excel 数据导入工具")
    print("=" * 50)

    init_db()

    # 导入新版库存表
    if os.path.exists('Stock_table.xlsx'):
        print(f"\n找到 Stock_table.xlsx，开始导入...")
        import_stock_table('Stock_table.xlsx')

    # 导入旧版库存表
    if os.path.exists('库存表.xlsx'):
        print(f"\n找到 库存表.xlsx，开始导入...")
        import_old_stock('库存表.xlsx')

    # 导入 ISBN 对照表
    if os.path.exists('ISBN-货号对照表.xlsx'):
        print(f"\n找到 ISBN-货号对照表.xlsx，开始导入...")
        import_old_stock('ISBN-货号对照表.xlsx')

    print("\n" + "=" * 50)
    print("  所有数据导入完成！")
    print("  运行 python app.py 启动系统")
    print("=" * 50)
