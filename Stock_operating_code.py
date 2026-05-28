
"""
@author = allenchow
@email = 32502439@stu.hzcu.edu.cn
@current_time = 13/12/2025 14:05
功能：库存管理系统（支持入库、出库、扫码检测功能）
支持Excel文件自动初始化、库存变动日志记录、条码索引优化提升效率
"""

import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog, messagebox
import time
from tqdm import tqdm

# -------------------------- 核心配置（可根据实际需求修改） --------------------------
EXCEL_FILE = "Stock_table.xlsx"  # 库存表路径
CHECKLIST_SHEET = "19-number"  # 核对表sheet名
STOCK_SHEET = "Stock"  # 库存主表sheet名
LOG_SHEET = "log"  # 日志sheet名

# 固定表头配置
LOG_HEADERS = ["操作时间", "条形码", "货号", "变动量", "操作类型", "操作前库存", "操作后库存", "操作状态", "备注"]
STOCK_HEADERS = ["条码/19码", "货号", "商品名称", "品牌", "尺码", "库存数量", "备注"]
CHECKLIST_HEADERS = ["条码/19码", "货号", "商品名称", "品牌", "尺码", "备注"]


def init_excel():
    """
    初始化Excel文件
    如果文件不存在/缺少sheet/缺少表头，自动创建并补充
    """
    # 检查文件是否存在，不存在则创建
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        # 创建核对表
        checklist = wb.active
        checklist.title = CHECKLIST_SHEET
        checklist.append(CHECKLIST_HEADERS)
        # 创建库存表
        stock = wb.create_sheet(STOCK_SHEET)
        stock.append(STOCK_HEADERS)
        # 创建日志表
        log = wb.create_sheet(LOG_SHEET)
        log.append(LOG_HEADERS)
        wb.save(EXCEL_FILE)
        wb.close()

    # 加载已有文件，检查并补充缺失的sheet
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 检查核对表
    if CHECKLIST_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(CHECKLIST_SHEET)
        sheet.append(CHECKLIST_HEADERS)

    # 检查库存表
    if STOCK_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(STOCK_SHEET)
        sheet.append(STOCK_HEADERS)

    # 检查日志表
    if LOG_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(LOG_SHEET)
        sheet.append(LOG_HEADERS)

    wb.save(EXCEL_FILE)
    wb.close()


def write_log(ws, barcode, sku, change, op_type, before, after, status, remark=""):
    """
    写入操作日志
    :param ws: 日志工作表对象
    :param barcode: 条形码
    :param sku: 货号
    :param change: 库存变动量（+1入库/-1出库/0检测）
    :param op_type: 操作类型（入库/出库/检测）
    :param before: 操作前库存
    :param after: 操作后库存
    :param status: 操作状态（成功/失败/存在/不存在）
    :param remark: 备注信息
    """
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        barcode,
        sku,
        change,
        op_type,
        before,
        after,
        status,
        remark
    ])


def build_index(ws, start_row):
    """
    构建条码->行号的内存索引，提升查询效率
    :param ws: 工作表对象
    :param start_row: 起始行号（跳过表头）
    :return: 索引字典 {barcode: row_num}
    """
    index = {}
    for row in range(start_row, ws.max_row + 1):
        barcode = ws.cell(row, 1).value
        if barcode:
            index[str(barcode)] = row
    return index


def find_in_checklist(ws, checklist_index, barcode):
    """
    通过索引快速查询核对表中的条码信息
    :param ws: 核对表工作表对象
    :param checklist_index: 核对表索引
    :param barcode: 要查询的条形码
    :return: 商品信息字典 / None（未找到）
    """
    row = checklist_index.get(str(barcode))
    if not row:
        return None

    return {
        "条码/19码": ws.cell(row, 1).value,
        "货号": ws.cell(row, 2).value,
        "商品名称": ws.cell(row, 3).value,
        "品牌": ws.cell(row, 4).value,
        "尺码": ws.cell(row, 5).value,
        "备注": ws.cell(row, 6).value
    }


def manual_input(barcode):
    """
    手动录入未知条码的商品信息（建档）
    :param barcode: 条形码
    :return: 商品信息字典 / None（录入取消）
    """
    print(f"\n条码 {barcode} 未在核对表中，开始建档，请依次输入信息：")

    # 货号（必填）
    sku = input("请输入货号: ").strip()
    if not sku:
        print("货号不能为空，建档取消")
        return None

    # 商品名称（必填）
    name = input("请输入商品名称: ").strip()
    if not name:
        print("商品名称不能为空，建档取消")
        return None

    # 品牌（必填）
    brand = input("请输入品牌: ").strip()
    if not brand:
        print("品牌不能为空，建档取消")
        return None

    # 尺码（必填）
    size = input("请输入尺码: ").strip()
    if not size:
        print("尺码不能为空，建档取消")
        return None

    # 构建商品信息
    data = {
        "条码/19码": barcode,
        "货号": sku,
        "商品名称": name,
        "品牌": brand,
        "尺码": size,
        "备注": "自建库"
    }

    print("建档完成！")
    return data


def update_stock(ws, stock_index, data, mode):
    """
    更新库存数量
    :param ws: 库存表工作表对象
    :param stock_index: 库存表索引
    :param data: 商品信息字典
    :param mode: 操作模式（'in' 入库 / 'out' 出库）
    :return: (操作前库存, 操作后库存, 状态码)
            状态码：success-成功 / not_exist-出库商品不存在 / no_stock-库存不足 / deleted-出库后库存归零已删除
    """
    barcode = str(data["条码/19码"])

    # 条码不在库存表中
    if barcode not in stock_index:
        if mode == "out":
            return None, None, "not_exist"  # 出库模式下未知条码直接返回

        # 入库模式下新增商品
        ws.append([
            data["条码/19码"],
            data["货号"],
            data["商品名称"],
            data["品牌"],
            data["尺码"],
            1,
            data["备注"]
        ])
        new_row = ws.max_row
        stock_index[barcode] = new_row
        return 0, 1, "success"

    # 条码已存在，更新库存
    row = stock_index[barcode]
    before = ws.cell(row, 6).value or 0  # 操作前库存（默认0）

    # 出库操作
    if mode == "out":
        if before <= 0:
            return before, before, "no_stock"  # 库存不足
        elif before == 1:
            # 出库后库存为0，删除该行并更新索引
            ws.delete_rows(row, 1)
            del stock_index[barcode]

            # 修正删除行后的索引（后续行号减1）
            for k, v in stock_index.items():
                if v > row:
                    stock_index[k] = v - 1

            return 1, 0, "deleted"
        else:
            # 正常出库（库存>1）
            ws.cell(row, 6, before - 1)
            return before, before - 1, "success"

    # 入库操作
    ws.cell(row, 6, before + 1)
    return before, before + 1, "success"


def continuous_scan(mode):
    """
    连续扫码操作
    :param mode: 操作模式（'in' 入库 / 'out' 出库 / 'check' 扫码检测）
    :return: 未知条码（入库模式下）/ None
    """
    # 加载Excel文件
    wb = load_workbook(EXCEL_FILE)
    checklist_ws = wb[CHECKLIST_SHEET]
    stock_ws = wb[STOCK_SHEET]
    log_ws = wb[LOG_SHEET]

    # 构建索引提升查询效率
    checklist_index = build_index(checklist_ws, 3)
    stock_index = build_index(stock_ws, 2)

    unknown_barcode = None
    total_count = 0
    exist_count = 0
    not_exist_count = 0
    not_exist_list = []

    # 打印模式提示
    if mode == "in":
        print("\n=== 入库模式 ===")
        print("扫码录入，输入 exit 退出并保存\n")
    elif mode == "out":
        print("\n=== 出库模式 ===")
        print("扫码录入，输入 exit 退出并保存\n")
    elif mode == "check":
        print("\n=== 扫码检测模式 ===")
        print("扫码检测条码是否在核对表中，输入 exit 退出\n")

    # 循环扫码
    while True:
        barcode = input("扫码：").strip()

        # 空输入跳过
        if not barcode:
            continue

        # 退出指令
        if barcode.lower() == "exit":
            break

        # 扫码检测模式
        if mode == "check":
            total_count += 1
            data = find_in_checklist(checklist_ws, checklist_index, barcode)

            if data:
                exist_count += 1
                status = "存在"
                print(f"✔ 第{total_count}个商品在核对表中")
            else:
                not_exist_count += 1
                not_exist_list.append(total_count)
                status = "不存在"
                print(f"⚠ 第{total_count}个商品未在核对表中")

            # 写入检测日志
            write_log(
                log_ws,
                barcode,
                data["货号"] if data else "-",
                0,
                "检测",
                "-",
                "-",
                status,
                f"第{total_count}个商品"
            )
            continue

        # 入库/出库模式 - 查询核对表
        data = find_in_checklist(checklist_ws, checklist_index, barcode)

        # 条码不存在
        if not data:
            if mode == "out":
                print(f"⚠ 出库模式不允许未知条码：{barcode}")
                continue
            unknown_barcode = barcode
            print(f"⚠ 未知条码 {barcode}，退出扫码进入建档")
            break

        # 更新库存
        before, after, status = update_stock(stock_ws, stock_index, data, mode)

        # 库存不足处理
        if status == "no_stock":
            print(f"⚠ 库存不足，无法出库：{barcode}")
            continue

        # 出库后删除库存行
        elif status == "deleted":
            write_log(
                log_ws,
                barcode,
                data["货号"],
                -1,
                "出库",
                before,
                after,
                "成功",
                "库存归零已删除"
            )
            print(f"✔ {barcode} 出库完成，库存归零，已从库存表删除")
            continue

        # 正常入库/出库
        write_log(
            log_ws,
            barcode,
            data["货号"],
            -1 if mode == "out" else 1,
            "出库" if mode == "out" else "入库",
            before,
            after,
            "成功"
        )
        print(f"✔ {barcode} {'出库' if mode == 'out' else '入库'}完成（{before} → {after}）")

    # 保存并关闭文件
    wb.save(EXCEL_FILE)
    wb.close()

    # 检测模式统计结果
    if mode == "check":
        print("\n=== 扫描统计结果 ===")
        print(f"共扫描 {total_count} 个商品")
        print(f"在核对表中的商品数：{exist_count}")
        print(f"未在核对表中的商品数：{not_exist_count}")
        if not_exist_list:
            print(f"未在核对表中的商品是第 {not_exist_list} 个商品")

    return unknown_barcode if mode == "in" else None


if __name__ == "__main__":
    # 初始化Excel文件
    init_excel()

    # 主程序循环
    while True:
        print("\n==================== 库存管理系统 ====================")
        print("请选择操作模式：")
        print("1 - 入库")
        print("2 - 出库")
        print("3 - 扫码检测模式")
        print("0 - 退出程序")
        print("=====================================================")

        choice = input("请输入 0/1/2/3：").strip()

        # 入库模式
        if choice == "1":
            unknown = continuous_scan(mode="in")
            # 处理未知条码建档
            if unknown:
                print(f"\n开始为条码 {unknown} 建档")
                data = manual_input(unknown)
                if data:
                    # 重新加载文件写入数据
                    wb = load_workbook(EXCEL_FILE)
                    checklist_ws = wb[CHECKLIST_SHEET]
                    stock_ws = wb[STOCK_SHEET]
                    log_ws = wb[LOG_SHEET]

                    # 写入核对表
                    checklist_ws.append([
                        data["条码/19码"],
                        data["货号"],
                        data["商品名称"],
                        data["品牌"],
                        data["尺码"],
                        data["备注"]
                    ])

                    # 写入库存表
                    stock_ws.append([
                        data["条码/19码"],
                        data["货号"],
                        data["商品名称"],
                        data["品牌"],
                        data["尺码"],
                        1,
                        data["备注"]
                    ])

                    # 写入日志
                    write_log(
                        log_ws,
                        unknown,
                        data["货号"],
                        1,
                        "自建入库",
                        0,
                        1,
                        "成功",
                        "自建库"
                    )

                    wb.save(EXCEL_FILE)
                    wb.close()
                    print(f"✔ 条码 {unknown} 建档并入库完成！")

        # 出库模式
        elif choice == "2":
            continuous_scan(mode="out")

        # 扫码检测模式
        elif choice == "3":
            continuous_scan(mode="check")

        # 退出程序
        elif choice == "0":
            print("\n程序已退出，感谢使用！")
            break

        # 无效选择
        else:
            print("\n❌ 无效选择，请输入 0/1/2/3 中的一个数字！")