"""
@author = allenchow
@email = 32502439@stu.hzcu.edu.cn
@current_time = 8/12/2025 16:32
"""
import openpyxl
from datetime import datetime
import os

#sheet文件说明
EXCEL_FILE = "库存表.xlsx"  # 库存表路径
MATCH_SHEET = "ISBN-货号对照表"  # 对照表sheet名
LOG_SHEET = "操作日志"  # 日志sheet名
# 日志表头（如果日志表为空，自动创建）
LOG_HEADERS = ["操作时间", "ISBN码", "货号", "变动量", "操作类型", "操作前库存", "操作后库存", "操作状态", "备注"]


# ===================== 工具函数 =====================
def init_excel():
    """初始化Excel文件（如果不存在/缺少sheet/缺少表头，自动创建）"""
    # 1. 如果文件不存在，新建文件
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        # 创建对照表sheet
        map_sheet = wb.active
        map_sheet.title = MATCH_SHEET
        map_sheet.append(["ISBN码", "货号", "当前库存"])  # 对照表表头
        # 创建日志sheet
        log_sheet = wb.create_sheet(LOG_SHEET)
        log_sheet.append(LOG_HEADERS)  # 日志表头
        wb.save(EXCEL_FILE)
        wb.close()
        print(f"首次运行：已创建{EXCEL_FILE}，包含{MATCH_SHEET }和{LOG_SHEET}")

    # 2. 检查日志sheet的表头是否完整
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if LOG_SHEET not in wb.sheetnames:
        log_sheet = wb.create_sheet(LOG_SHEET)
        log_sheet.append(LOG_HEADERS)
    else:
        log_sheet = wb[LOG_SHEET]
        # 检查第一行是否是标准表头
        first_row = [cell.value for cell in log_sheet[1]]
        if first_row != LOG_HEADERS:
            log_sheet.insert_rows(1)  # 插入新行放表头
            for col, header in enumerate(LOG_HEADERS, 1):
                log_sheet.cell(row=1, column=col, value=header)
    wb.save(EXCEL_FILE)
    wb.close()


def write_log(log_data):
    """写入日志（通用函数，接收日志字典，自动追加到日志表最后一行）"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    log_sheet = wb[LOG_SHEET]

    # 找到日志表的最后一行（跳过空行）
    last_row = 1
    for row in log_sheet.iter_rows(min_row=1, values_only=True):
        if any(cell for cell in row):  # 只要行中有非空值，就算有效行
            last_row += 1

    # 按表头顺序写入日志数据
    for col, header in enumerate(LOG_HEADERS, 1):
        log_sheet.cell(row=last_row, column=col, value=log_data.get(header, ""))

    wb.save(EXCEL_FILE)
    wb.close()


# ===================== 主逻辑 =====================
if __name__ == "__main__":
    # 1. 初始化Excel（确保文件/表头存在）
    init_excel()

    # 2. 打开Excel文件
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet_map = wb[MATCH_SHEET ]
    except PermissionError:
        print(f"错误：请先关闭Excel文件「{EXCEL_FILE}」，再运行代码！")
        exit()

    # 3. 构建ISBN→货号/库存/行号的字典
    isbn_dict = {}
    for row in sheet_map.iter_rows(min_row=2):  # 从第2行开始（跳过表头）
        cell_isbn = row[0]
        isbn_code = cell_isbn.value
        goods_no = row[1].value if row[1].value else ""
        stock = row[2].value if row[2].value else 0  # 空库存默认0

        if isbn_code:  # 跳过空ISBN行
            isbn_dict[isbn_code] = {
                "货号": goods_no,
                "库存": stock,
                "行号": cell_isbn.row
            }

    # 4. 读取扫码的ISBN
    isbn = input("请扫码输入ISBN（或手动输入）：").strip()
    # 初始化日志数据（默认操作失败）
    log_data = {
        "操作时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],  # 精确到毫秒
        "ISBN码": isbn,
        "货号": "",
        "变动量": "",
        "操作类型": "",
        "操作前库存": "",
        "操作后库存": "",
        "操作状态": "失败",
        "备注": ""
    }

    # 5. 匹配货号并处理库存
    if not isbn:
        log_data["备注"] = "未输入ISBN码，操作终止"
        write_log(log_data)
        print("错误：未输入ISBN码！")
    elif isbn not in isbn_dict:
        log_data["备注"] = f"未找到ISBN「{isbn}」对应的货号"
        write_log(log_data)
        print(f"错误：未找到ISBN {isbn} 对应的货号！")
    else:
        goods_info = isbn_dict[isbn]
        log_data["货号"] = goods_info["货号"]
        log_data["操作前库存"] = goods_info["库存"]

        # 输入变动量并校验
        try:
            change_num = int(
                input(f"货号「{goods_info['货号']}」，当前库存「{goods_info['库存']}」，请输入变动量（+入库/-出库）："))
            log_data["变动量"] = change_num

            # 判断操作类型
            if change_num > 0:
                log_data["操作类型"] = "入库"
            elif change_num < 0:
                log_data["操作类型"] = "出库"
            else:
                log_data["操作类型"] = "无变动"
                log_data["备注"] = "变动量为0，未修改库存"
                write_log(log_data)
                print("提示：变动量为0，无需修改库存！")
                wb.close()
                exit()

            # 计算新库存并更新
            new_stock = goods_info["库存"] + change_num
            sheet_map.cell(row=goods_info["行号"], column=3, value=new_stock)

            # 更新日志为成功
            log_data["操作后库存"] = new_stock
            log_data["操作状态"] = "成功"
            log_data["备注"] = f"{log_data['操作类型']}{abs(change_num)}件"

            print(f"操作成功！货号「{goods_info['货号']}」，{log_data['操作类型']}{abs(change_num)}件，新库存：{new_stock}")
        except ValueError:
            log_data["备注"] = "输入的变动量不是有效数字（需输入整数，如1、-2）"
            write_log(log_data)
            print("错误：请输入数字（如1、-2）！")
        except Exception as e:
            log_data["备注"] = f"系统异常：{str(e)}"
            write_log(log_data)
            print(f"错误：系统异常 - {str(e)}")

    # 6. 保存并关闭文件
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        print(f"错误：请先关闭Excel文件「{EXCEL_FILE}」，再运行代码！")
    finally:
        wb.close()

    # 7. 确保日志写入（即使库存操作失败也会记录）
    if "log_data" in locals():
        write_log(log_data)